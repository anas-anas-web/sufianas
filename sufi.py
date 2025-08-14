import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ==================== KONSTANTLAR ====================
VOWELS = 'aıouəeiöü'
CONSONANTS = 'bcçdfgğhjklmnprsştvxyz'
BACK_VOWELS = 'aıou'
FRONT_VOWELS = 'əeiöü'

SPECIAL_WORDS = {
    'su': {
        'plural': 'sular',
        'possessive': {'1s': 'suyum', '3s': 'suyu'},
        'case': {'Yönlük': 'suya', 'Yerlik': 'suda'}
    },
    'ata': {'case': {'Yönlük': 'ataya'}},
    'ana': {'possessive': {'1s': 'anam'}}
}

# ==================== KÖMƏKÇİ FUNKSİYALAR ====================
def get_last_vowel(word):
    for ch in reversed(word):
        if ch in VOWELS:
            return ch
    return None

def detect_pos(word):
    parts_of_speech = {
        "pronouns": {"mən", "sən", "o", "biz", "siz", "onlar"},
        "numerals": {"bir", "iki", "üç", "dörd", "beş", "altı", "yeddi", "səkkiz", "doqquz", "on"},
        "particles": {"də", "belə", "yalnız"},
        "conjunctions": {"və", "ya", "amma", "çünki"},
        "prepositions": {"ilə", "üçün", "qarşı", "üstə"},
        "interjections": {"əə", "ay", "vau"}
    }
    if word in parts_of_speech["pronouns"]: return "Əvəzlik"
    if word in parts_of_speech["numerals"]: return "Say"
    if word in parts_of_speech["particles"]: return "Hərf"
    if word in parts_of_speech["conjunctions"]: return "Qoşma"
    if word in parts_of_speech["prepositions"]: return "Ədat"
    if word in parts_of_speech["interjections"]: return "Səslər"
    if word.endswith(("maq", "mək")): return "Fel"
    if word.endswith(("ca", "cə", "la", "lə")): return "Zərf"
    if word.endswith(("lı", "li", "lu", "lü")): return "Sifət"
    return "İsim"

def generate_plural(word):
    if word in SPECIAL_WORDS:
        return SPECIAL_WORDS[word].get('plural', word+'lar')
    last_v = get_last_vowel(word)
    if not last_v:
        return word
    if last_v in BACK_VOWELS:
        return f"{word}lar"
    return f"{word}lər"

def generate_case(word, case):
    if word in SPECIAL_WORDS and case in SPECIAL_WORDS[word].get('case', {}):
        return SPECIAL_WORDS[word]['case'][case]
    last_v = get_last_vowel(word)
    if not last_v:
        return word
    # Sonu samitlə bitənlər
    if word[-1] not in VOWELS:
        case_map = {
            'a': {'Yiyəlik': 'ın', 'Yönlük': 'a', 'Təsirlik': 'ı', 'Yerlik': 'da', 'Çıxışlıq': 'dan'},
            'ı': {'Yiyəlik': 'ın', 'Yönlük': 'a', 'Təsirlik': 'ı', 'Yerlik': 'da', 'Çıxışlıq': 'dan'},
            'ə': {'Yiyəlik': 'in', 'Yönlük': 'ə', 'Təsirlik': 'i', 'Yerlik': 'də', 'Çıxışlıq': 'dən'},
            'i': {'Yiyəlik': 'in', 'Yönlük': 'ə', 'Təsirlik': 'i', 'Yerlik': 'də', 'Çıxışlıq': 'dən'},
            'o': {'Yiyəlik': 'un', 'Yönlük': 'a', 'Təsirlik': 'u', 'Yerlik': 'da', 'Çıxışlıq': 'dan'},
            'u': {'Yiyəlik': 'un', 'Yönlük': 'a', 'Təsirlik': 'u', 'Yerlik': 'da', 'Çıxışlıq': 'dan'},
            'ö': {'Yiyəlik': 'ün', 'Yönlük': 'ə', 'Təsirlik': 'ü', 'Yerlik': 'də', 'Çıxışlıq': 'dən'},
            'ü': {'Yiyəlik': 'ün', 'Yönlük': 'ə', 'Təsirlik': 'ü', 'Yerlik': 'də', 'Çıxışlıq': 'dən'}
        }
        suffix = case_map.get(last_v, {}).get(case, '')
        return f"{word}{suffix}"
    # Sonu saitlə bitənlər
    else:
        case_map = {
            'a': {'Yiyəlik': 'nın', 'Yönlük': 'ya', 'Təsirlik': 'nı', 'Yerlik': 'da', 'Çıxışlıq': 'dan'},
            'ı': {'Yiyəlik': 'nın', 'Yönlük': 'ya', 'Təsirlik': 'nı', 'Yerlik': 'da', 'Çıxışlıq': 'dan'},
            'ə': {'Yiyəlik': 'nin', 'Yönlük': 'yə', 'Təsirlik': 'ni', 'Yerlik': 'də', 'Çıxışlıq': 'dən'},
            'i': {'Yiyəlik': 'nin', 'Yönlük': 'yə', 'Təsirlik': 'ni', 'Yerlik': 'də', 'Çıxışlıq': 'dən'},
            'o': {'Yiyəlik': 'nun', 'Yönlük': 'ya', 'Təsirlik': 'nu', 'Yerlik': 'da', 'Çıxışlıq': 'dan'},
            'u': {'Yiyəlik': 'nun', 'Yönlük': 'ya', 'Təsirlik': 'nu', 'Yerlik': 'da', 'Çıxışlıq': 'dan'},
            'ö': {'Yiyəlik': 'nün', 'Yönlük': 'yə', 'Təsirlik': 'nü', 'Yerlik': 'də', 'Çıxışlıq': 'dən'},
            'ü': {'Yiyəlik': 'nün', 'Yönlük': 'yə', 'Təsirlik': 'nü', 'Yerlik': 'də', 'Çıxışlıq': 'dən'}
        }
        suffix = case_map.get(last_v, {}).get(case, '')
        return f"{word}{suffix}"

def generate_possessive(word, person="1s", plural=False):
    if word in SPECIAL_WORDS and person in SPECIAL_WORDS[word].get('possessive', {}):
        return SPECIAL_WORDS[word]['possessive'][person]
    last_v = get_last_vowel(word)
    if not last_v:
        return word
    base = generate_plural(word) if plural else word
    # Vowel-ending words
    if word[-1] in VOWELS:
        suffix_map = {
            "1s": {"a": "m", "ı": "m", "ə": "m", "i": "m", "o": "m", "u": "m", "ö": "m", "ü": "m"},
            "2s": {"a": "n", "ı": "n", "ə": "n", "i": "n", "o": "n", "u": "n", "ö": "n", "ü": "n"},
            "3s": {"a": "sı", "ı": "sı", "ə": "si", "i": "si", "o": "su", "u": "su", "ö": "sü", "ü": "sü"},
            "1p": {"a": "mız", "ı": "mız", "ə": "miz", "i": "miz", "o": "muz", "u": "muz", "ö": "müz", "ü": "müz"},
            "2p": {"a": "nız", "ı": "nız", "ə": "niz", "i": "niz", "o": "nuz", "u": "nuz", "ö": "nüz", "ü": "nüz"},
            "3p": {"a": "ları", "ı": "ları", "ə": "ləri", "i": "ləri", "o": "ları", "u": "ları", "ö": "ləri", "ü": "ləri"}
        }
        suffix = suffix_map.get(person, {}).get(last_v, "")
        return f"{base}{suffix}"
    # Consonant-ending words
    else:
        suffix_map = {
            "1s": {"a": "ım", "ı": "ım", "ə": "im", "i": "im", "o": "um", "u": "um", "ö": "üm", "ü": "üm"},
            "2s": {"a": "ın", "ı": "ın", "ə": "in", "i": "in", "o": "un", "u": "un", "ö": "ün", "ü": "ün"},
            "3s": {"a": "ı", "ı": "ı", "ə": "i", "i": "i", "o": "u", "u": "u", "ö": "ü", "ü": "ü"},
            "1p": {"a": "ımız", "ı": "ımız", "ə": "imiz", "i": "imiz", "o": "umuz", "u": "umuz", "ö": "ümüz", "ü": "ümüz"},
            "2p": {"a": "ınız", "ı": "ınız", "ə": "iniz", "i": "iniz", "o": "unuz", "u": "unuz", "ö": "ünüz", "ü": "ünüz"},
            "3p": {"a": "ları", "ı": "ları", "ə": "ləri", "i": "ləri", "o": "ları", "u": "ları", "ö": "ləri", "ü": "ləri"}
        }
        suffix = suffix_map.get(person, {}).get(last_v, "")
        return f"{base}{suffix}"

def generate_xeberlik(word, person="3s"):
    last_v = get_last_vowel(word)
    if not last_v:
        return word
    # Vowel-ending words
    if word[-1] in VOWELS:
        suffix_map = {
            "1s": {"a": "yam", "ı": "yam", "ə": "yəm", "i": "yəm", "o": "yam", "u": "yam", "ö": "yəm", "ü": "yəm"},
            "2s": {"a": "san", "ı": "san", "ə": "sən", "i": "sən", "o": "san", "u": "san", "ö": "sən", "ü": "sən"},
            "3s": {"a": "dır", "ı": "dır", "ə": "dir", "i": "dir", "o": "dur", "u": "dur", "ö": "dür", "ü": "dür"},
            "1p": {"a": "yıq", "ı": "yıq", "ə": "yik", "i": "yik", "o": "yuq", "u": "yuq", "ö": "yük", "ü": "yük"},
            "2p": {"a": "sınız", "ı": "sınız", "ə": "siniz", "i": "siniz", "o": "sunuz", "u": "sunuz", "ö": "sünüz", "ü": "sünüz"},
            "3p": {"a": "dırlar", "ı": "dırlar", "ə": "dirlər", "i": "dirlər", "o": "durlar", "u": "durlar", "ö": "dürlər", "ü": "dürlər"}
        }
        suffix = suffix_map.get(person, {}).get(last_v, "")
        return f"{word}{suffix}"
    # Consonant-ending words
    else:
        suffix_map = {
            "1s": {"a": "am", "ı": "am", "ə": "əm", "i": "əm", "o": "am", "u": "am", "ö": "əm", "ü": "əm"},
            "2s": {"a": "san", "ı": "san", "ə": "sən", "i": "sən", "o": "san", "u": "san", "ö": "sən", "ü": "sən"},
            "3s": {"a": "dır", "ı": "dır", "ə": "dir", "i": "dir", "o": "dur", "u": "dur", "ö": "dür", "ü": "dür"},
            "1p": {"a": "ıq", "ı": "ıq", "ə": "ik", "i": "ik", "o": "uq", "u": "uq", "ö": "ük", "ü": "ük"},
            "2p": {"a": "sınız", "ı": "sınız", "ə": "siniz", "i": "siniz", "o": "sunuz", "u": "sunuz", "ö": "sünüz", "ü": "sünüz"},
            "3p": {"a": "dırlar", "ı": "dırlar", "ə": "dirlər", "i": "dirlər", "o": "durlar", "u": "durlar", "ö": "dürlər", "ü": "dürlər"}
        }
        suffix = suffix_map.get(person, {}).get(last_v, "")
        return f"{word}{suffix}"

# ==================== ƏSAS EMAL FUNKSİYASI ====================
def process_words(input_file, output_file, suffixes_file=None):
    df_input = pd.read_excel(input_file)
    words = df_input['Söz'].astype(str).tolist()

    suffix_examples = {}
    if suffixes_file:
        suffix_examples = read_suffix_examples_from_excel(suffixes_file)

    results = {
        'Bütün_Sözlər': [],
        'Cəm_Formaları': [],
        'Hal_Şəkilçiləri': [],
        'Mənsubiyyət_Şəkilçiləri': [],
        'Xəbərlik_Şəkilçiləri': []
    }

    cases = ['Adlıq', 'Yiyəlik', 'Yönlük', 'Təsirlik', 'Yerlik', 'Çıxışlıq']
    persons = ['1s', '2s', '3s', '1p', '2p', '3p']
    person_names = ['mən', 'sən', 'o', 'biz', 'siz', 'onlar']

    for word in words:
        # Cəm forması
        plural = generate_plural(word)
        suffix = plural.replace(word, '')
        plural_with_suffix = f"{word}+{suffix}" if suffix else word
        example = suffix_examples.get(suffix, '')
        plural_with_example = f"{plural_with_suffix} ({example})" if example else plural_with_suffix
        results['Cəm_Formaları'].append({'Söz': word, 'Cəm forması': plural_with_example})
        results['Bütün_Sözlər'].append({'Yeni Söz': plural_with_example})

        # Hal şəkilçiləri
        case_data = {'Söz': word}
        for case in cases:
            form = generate_case(word, case)
            suffix = form.replace(word, '')
            case_form = f"{word}+{suffix}" if suffix else word
            example = suffix_examples.get(suffix, '')
            case_form_with_example = f"{case_form} ({example})" if example else case_form
            case_data[case] = case_form_with_example
            results['Bütün_Sözlər'].append({'Yeni Söz': case_form_with_example})
        results['Hal_Şəkilçiləri'].append(case_data)

        # Mənsubiyyət şəkilçiləri
        poss_data = {'Söz': word}
        for p in persons:
            tek_form = generate_possessive(word, p)
            tek_suffix = tek_form.replace(word, '')
            tek_word = f"{word}+{tek_suffix}" if tek_suffix else word
            example = suffix_examples.get(tek_suffix, '')
            tek_word_with_example = f"{tek_word} ({example})" if example else tek_word
            poss_data[p + "_tək"] = tek_word_with_example
            results['Bütün_Sözlər'].append({'Yeni Söz': tek_word_with_example})

            plural_root = generate_plural(word)
            cem_form = generate_possessive(word, p, plural=True)
            cem_suffix = cem_form.replace(plural_root, '')
            cem_word = f"{plural_root}+{cem_suffix}" if cem_suffix else plural_root
            example = suffix_examples.get(cem_suffix, '')
            cem_word_with_example = f"{cem_word} ({example})" if example else cem_word
            poss_data[p + "_cəm"] = cem_word_with_example
            results['Bütün_Sözlər'].append({'Yeni Söz': cem_word_with_example})
        results['Mənsubiyyət_Şəkilçiləri'].append(poss_data)

        # Xəbərlik şəkilçiləri
        x_data = {'Söz': word}
        for p, name in zip(persons, person_names):
            form = generate_xeberlik(word, p)
            suffix = form.replace(word, '')
            xeberlik_form = f"{word}+{suffix}" if suffix else word
            example = suffix_examples.get(suffix, '')
            xeberlik_form_with_example = f"{xeberlik_form} ({example})" if example else xeberlik_form
            x_data[name] = xeberlik_form_with_example
            results['Bütün_Sözlər'].append({'Yeni Söz': xeberlik_form_with_example})
        results['Xəbərlik_Şəkilçiləri'].append(x_data)

    # Excel-ə yazma
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        pd.DataFrame(results['Bütün_Sözlər']).to_excel(writer, sheet_name='Bütün_Sözlər', index=False)
        for sheet in ['Cəm_Formaları', 'Hal_Şəkilçiləri', 'Mənsubiyyət_Şəkilçiləri', 'Xəbərlik_Şəkilçiləri']:
            pd.DataFrame(results[sheet]).to_excel(writer, sheet_name=sheet, index=False)

    wb = load_workbook(output_file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
    wb.save(output_file)
    print(f"✅ Excel faylı '{output_file}' uğurla yaradıldı!")

# ==================== ŞƏKİLÇİLƏR VƏ NÜMUNƏLƏRİNİ ÇIXARAN FUNKSİYA ====================
def extract_unique_suffixes_and_examples_with_code_suffixes(file_path, output_path):
    wb = load_workbook(file_path)
    ws = wb['Bütün_Sözlər']

    suffix_dict = {}

    # 1. Bütün_Sözlər sheet-indən şəkilçiləri çıxart
    for row in ws.iter_rows(min_row=2, values_only=True):
        cell = row[0]
        if cell and '+' in cell:
            parts = str(cell).split('+')
            if len(parts) == 2:
                root, suffix = parts
                suffix = suffix.strip()
                if suffix not in suffix_dict:
                    suffix_dict[suffix] = cell  # ilk nümunə

    # 2. Koddan şəkilçilər əlavə et
    CODE_SUFFIXES = [
        'lar', 'lər', 'ın', 'in', 'a', 'ə', 'ı', 'i', 'da', 'də', 'dan', 'dən',
        'm', 'n', 'sı', 'si', 'su', 'sü', 'ım', 'im', 'um', 'üm',
        'ımız', 'imiz', 'umuz', 'ümüz', 'ınız', 'iniz', 'unuz', 'ünüz',
        'ları', 'ləri',
        'yam', 'yəm', 'san', 'sən', 'dır', 'dir', 'dur', 'dür',
        'ıq', 'ik', 'uq', 'ük', 'sınız', 'siniz', 'sunuz', 'sünüz',
        'dırlar', 'dirlər', 'durlar', 'dürlər'
    ]
    for suffix in CODE_SUFFIXES:
        if suffix not in suffix_dict:
            suffix_dict[suffix] = ''
    sorted_suffixes = sorted(suffix_dict.items(), key=lambda x: x[0])
    df = pd.DataFrame(sorted_suffixes, columns=["Şəkilçi", "Nümunə"])
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name="Şəkilçilər və Nümunələr", index=False)
   # print("✅ Kodda olan və Excel-də olmayan şəkilçilər də əlavə olundu!")

# ==================== İSİMLƏR SHEETİNİ YARAT VƏ QURUPLAŞDIRDIQ ====================
def add_grouped_headers_to_isimler_sheet(file_path):
    wb = load_workbook(file_path)
    ws = wb['İsimlər']
    old_headers = [cell.value for cell in ws[1]]
    hal_headers = ['Adlıq', 'Yiyəlik', 'Yönlük', 'Təsirlik', 'Yerlik', 'Çıxışlıq']
    mens_headers = [
        '1s_tək', '1s_cəm', '2s_tək', '2s_cəm', '3s_tək', '3s_cəm',
        '1p_tək', '1p_cəm', '2p_tək', '2p_cəm', '3p_tək', '3p_cəm'
    ]
    xeberlik_headers = ['Söz', 'mən', 'sən', 'o', 'biz', 'siz', 'onlar']

    def find_indices(header_list):
        return [i+1 for i, h in enumerate(old_headers) if h in header_list and i >= 2]

    hal_idx = find_indices(hal_headers)
    mens_idx = find_indices(mens_headers)
    xeb_idx = find_indices(xeberlik_headers)

    new_header = [''] * len(old_headers)
    if hal_idx:
        new_header[hal_idx[0]-1] = 'Hal_Şəkilçiləri'
    if mens_idx:
        new_header[mens_idx[0]-1] = 'Mənsubiyyət_Şəkilçiləri'
    if xeb_idx:
        new_header[xeb_idx[0]-1] = 'Xəbərlik_Şəkilçiləri'

    ws.insert_rows(1)
    for col, val in enumerate(new_header, 1):
        ws.cell(row=1, column=col, value=val)

    if hal_idx and len(hal_idx) > 1:
        ws.merge_cells(start_row=1, start_column=hal_idx[0], end_row=1, end_column=hal_idx[-1])
    if mens_idx and len(mens_idx) > 1:
        ws.merge_cells(start_row=1, start_column=mens_idx[0], end_row=1, end_column=mens_idx[-1])
    if xeb_idx and len(xeb_idx) > 1:
        ws.merge_cells(start_row=1, start_column=xeb_idx[0], end_row=1, end_column=xeb_idx[-1])

    for col in hal_idx + mens_idx + xeb_idx:
        cell = ws.cell(row=1, column=col)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    wb.save(file_path)
   # print("✅ İsimlər sheet-inə qruplaşdırılmış başlıqlar əlavə olundu!")

def color_grouped_headers_in_isimler(file_path):
    wb = load_workbook(file_path)
    ws = wb['İsimlər']
    # Qrup başlıqları və rəngləri
    colors = {
        'Hal_Şəkilçiləri': 'FFD966',        # Sarı
        'Mənsubiyyət_Şəkilçiləri': 'A9D08E',# Yaşıl
        'Xəbərlik_Şəkilçiləri': '9DC3E6'    # Mavi
    }
    # Alt başlıqlar və uyğun rəng
    hal_headers = ['Adlıq', 'Yiyəlik', 'Yönlük', 'Təsirlik', 'Yerlik', 'Çıxışlıq']
    mens_headers = [
        '1s_tək', '1s_cəm', '2s_tək', '2s_cəm', '3s_tək', '3s_cəm',
        '1p_tək', '1p_cəm', '2p_tək', '2p_cəm', '3p_tək', '3p_cəm'
    ]
    xeberlik_headers = ['mən', 'sən', 'o', 'biz', 'siz', 'onlar']

    for col in range(1, ws.max_column + 1):
        val1 = ws.cell(row=1, column=col).value
        val2 = ws.cell(row=2, column=col).value
        # Qrup başlıqları üçün
        if val1 in colors:
            merged_range = None
            for m in ws.merged_cells.ranges:
                if m.min_col == col and m.min_row == 1:
                    merged_range = m
                    break
            if merged_range:
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    cell = ws.cell(row=1, column=c)
                    cell.fill = PatternFill(start_color=colors[val1], end_color=colors[val1], fill_type="solid")
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell = ws.cell(row=1, column=col)
                cell.fill = PatternFill(start_color=colors[val1], end_color=colors[val1], fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        # Alt başlıqlar üçün
        if val2 in hal_headers:
            color = colors['Hal_Şəkilçiləri']
        elif val2 in mens_headers:
            color = colors['Mənsubiyyət_Şəkilçiləri']
        elif val2 in xeberlik_headers:
            color = colors['Xəbərlik_Şəkilçiləri']
        else:
            color = None
        if color:
            cell = ws.cell(row=2, column=col)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(file_path)
   # print("✅ İsimlər sheet-ində başlıqlar rəngləndi və səliqəyə salındı.")

def color_multiindex_headers(ws):
    """MultiIndex başlıqlarına və alt başlıqlara rəngli fon verir."""
    # Qrup başlıqları üçün rənglər
    group_colors = {
        'Hal_Şəkilçiləri': 'FFD966',        # Sarı
        'Mənsubiyyət_Şəkilçiləri': 'A9D08E',# Yaşıl
        'Xəbərlik_Şəkilçiləri': '9DC3E6'    # Mavi
    }
    # Alt başlıqlar üçün uyğunluq
    hal_headers = ['Adlıq', 'Yiyəlik', 'Yönlük', 'Təsirlik', 'Yerlik', 'Çıxışlıq']
    mens_headers = [
        '1s_tək', '1s_cəm', '2s_tək', '2s_cəm', '3s_tək', '3s_cəm',
        '1p_tək', '1p_cəm', '2p_tək', '2p_cəm', '3p_tək', '3p_cəm'
    ]
    xeb_headers = ['mən', 'sən', 'o', 'biz', 'siz', 'onlar']

    # 1-ci sətr: Qrup başlıqları
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val in group_colors:
            cell = ws.cell(row=1, column=col)
            cell.fill = PatternFill(start_color=group_colors[val], end_color=group_colors[val], fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 2-ci sətr: Alt başlıqlar
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val in hal_headers:
            color = group_colors['Hal_Şəkilçiləri']
        elif val in mens_headers:
            color = group_colors['Mənsubiyyət_Şəkilçiləri']
        elif val in xeb_headers:
            color = group_colors['Xəbərlik_Şəkilçiləri']
        else:
            color = None
        cell = ws.cell(row=2, column=col)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

def create_isimler_sheet_with_grouped_headers(output_file):
    df_hal = pd.read_excel(output_file, sheet_name="Hal_Şəkilçiləri")
    df_mens = pd.read_excel(output_file, sheet_name="Mənsubiyyət_Şəkilçiləri")
    df_xeb = pd.read_excel(output_file, sheet_name="Xəbərlik_Şəkilçiləri")
    base_cols = ['Söz']

    hal_cols = ['Adlıq', 'Yiyəlik', 'Yönlük', 'Təsirlik', 'Yerlik', 'Çıxışlıq']
    mens_cols = [
        '1s_tək', '1s_cəm', '2s_tək', '2s_cəm', '3s_tək', '3s_cəm',
        '1p_tək', '1p_cəm', '2p_tək', '2p_cəm', '3p_tək', '3p_cəm'
    ]
    xeb_cols = ['mən', 'sən', 'o', 'biz', 'siz', 'onlar']

    df = df_hal[base_cols].copy()
    for col in hal_cols:
        df[col] = df_hal[col]
    for col in mens_cols:
        df[col] = df_mens[col]
    for col in xeb_cols:
        df[col] = df_xeb[col]

    arrays = [
        ['Söz'] + ['Hal_Şəkilçiləri']*len(hal_cols) + ['Mənsubiyyət_Şəkilçiləri']*len(mens_cols) + ['Xəbərlik_Şəkilçiləri']*len(xeb_cols),
        ['Söz'] + hal_cols + mens_cols + xeb_cols
    ]
    df.columns = pd.MultiIndex.from_arrays(arrays)

    write_multiindex_to_excel(df, output_file, "İsimlər")

    wb = load_workbook(output_file)
    remove_sheets(wb, ["Hal_Şəkilçiləri", "Mənsubiyyət_Şəkilçiləri", "Xəbərlik_Şəkilçiləri"])
    wb.save(output_file)

    ws = wb["İsimlər"]
    set_column_width_and_wrap(ws)
    color_multiindex_headers(ws)  # <-- Rəngləmə funksiyasını çağırın
    wb.save(output_file)

# ==================== QALAN FUNKSİYALAR EYNİ QALIR ====================
def set_column_width_and_wrap(ws, min_width=12, max_width=40):
    """Sütun enini və wrap_text-i tənzimləyir."""
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True,
                horizontal=cell.alignment.horizontal or 'left',
                vertical=cell.alignment.vertical or 'center'
            )
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(min_width, min(max_length + 2, max_width))

def remove_sheets(wb, sheet_names):
    """Verilmiş sheet-ləri silir."""
    for sheet in sheet_names:
        if sheet in wb.sheetnames:
            del wb[sheet]

def write_multiindex_to_excel(df, output_file, sheet_name):
    """MultiIndex başlıqlı DataFrame-i Excel-ə yazır."""
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name)  # index parametri olmadan, çünki MultiIndex-də index=False dəstəklənmir

def read_suffix_examples_from_excel(file_path):
    try:
        df = pd.read_excel(file_path, sheet_name="Şəkilçilər və Nümunələr")
        examples = {}
        for _, row in df.iterrows():
            suffix = str(row.get('Şəkilçi')).strip()
            example = str(row.get('Nümunə')).strip()
            if suffix and example:
                examples[suffix] = example
        return examples
    except Exception as e:
        print(f"Şəkilçilər və Nümunələr sheet-i tapılmadı və ya oxunmadı: {e}")
        return {}

# ==================== ƏSAS BLOK ====================
if __name__ == "__main__":
    process_words("input_2427.xlsx", "az_grammar_output.xlsx")
    extract_unique_suffixes_and_examples_with_code_suffixes("az_grammar_output.xlsx", "az_grammar_output.xlsx")
    create_isimler_sheet_with_grouped_headers("az_grammar_output.xlsx")
